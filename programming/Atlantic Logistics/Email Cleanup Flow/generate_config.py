"""
Email Cleanup Flow — Config File Generator
==========================================
Run this script once to produce email_cleanup_config.xlsx.
Upload that file to OneDrive and point the Power Automate flows at it.

To regenerate with fresh defaults: just run this script again.
It will overwrite the existing file.

Requires: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT_FILE = "email_cleanup_config.xlsx"

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
COLOR_HEADER_BLUE   = "1F4E79"   # dark blue — header rows
COLOR_HEADER_TEXT   = "FFFFFF"   # white text on headers
COLOR_SECTION_GRAY  = "D9D9D9"   # light gray — section labels
COLOR_DISABLED      = "F2F2F2"   # very light gray — disabled rows
COLOR_WARNING       = "FFF2CC"   # yellow — warning/caution cells
COLOR_SAFE          = "E2EFDA"   # light green — safe/enabled rows

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def header_style(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color=COLOR_HEADER_TEXT, size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def section_label(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill("solid", fgColor=COLOR_SECTION_GRAY)

def note_style(cell):
    cell.font = Font(italic=True, color="595959", size=9)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def thin_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------------------------------------------------------------------
# Sheet 1: RULES
# ---------------------------------------------------------------------------

def build_rules_sheet(wb):
    ws = wb.create_sheet("Rules")
    ws.sheet_view.showGridLines = True

    # --- Column headers ---
    headers = [
        "RuleName",
        "Enabled",
        "FromAddress",
        "SubjectContains",
        "BodyContains",
        "MinAge",
        "Action",
        "SourceFolder",
        "DestinationFolder",
        "Notes",
    ]

    col_widths = [22, 10, 32, 22, 22, 14, 12, 22, 22, 40]

    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=i)
        header_style(cell, h)
        set_col_width(ws, get_column_letter(i), w)

    ws.row_dimensions[1].height = 30

    # --- Example rules ---
    # Format: RuleName, Enabled, FromAddress, SubjectContains, BodyContains,
    #         MinAge, Action, SourceFolder, DestinationFolder, Notes
    # Use "none" for fields that don't apply to a rule.
    rules = [
        # Chained example: move Costco emails to archive folder after 7 days,
        # then queue them for deletion after 30 more days (37 days total).
        # Row 1 reads from Inbox. Row 2 reads from the Promotions folder.
        (
            "Costco (1) Archive",
            "yes",
            "costco@costco.com",
            "none",
            "none",
            "7d",
            "move",
            "none",
            "Promotions",
            "Step 1 of 2 - move Costco emails out of inbox into Promotions after 7 days.",
        ),
        (
            "Costco (2) Pending Delete",
            "yes",
            "costco@costco.com",
            "none",
            "none",
            "30d",
            "pending-delete",
            "Promotions",
            "Pending Delete",
            "Step 2 of 2 - queue Costco emails from Promotions for deletion after 30 days.",
        ),
        (
            "Automated Notifications",
            "yes",
            "noreply@example.com",
            "none",
            "none",
            "48h",
            "pending-delete",
            "none",
            "Pending Delete",
            "EXAMPLE - replace sender with actual noreply address. 48h min age.",
        ),
        (
            "Old Read Receipts",
            "no",
            "none",
            "Read:",
            "none",
            "1d",
            "pending-delete",
            "none",
            "Pending Delete",
            "DISABLED - example only. Enable and adjust if needed.",
        ),
    ]

    for row_num, rule in enumerate(rules, start=2):
        for col_num, value in enumerate(rule, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col_num == 10))
            # Color enabled vs disabled rows
            if rule[1].lower() == "yes":
                cell.fill = PatternFill("solid", fgColor=COLOR_SAFE)
                # "none" placeholder values — gray italic, visually non-values
                if value == "none":
                    cell.font = Font(italic=True, color="AAAAAA", size=9)
            else:
                cell.fill = PatternFill("solid", fgColor=COLOR_DISABLED)
                cell.font = Font(color="888888")

    # --- Excel Table ---
    last_row = 1 + len(rules)
    last_col = get_column_letter(len(headers))
    table = Table(displayName="RulesTable", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # --- Notes block below table ---
    notes_start = last_row + 2
    ws.cell(row=notes_start, column=1, value="FIELD REFERENCE").font = Font(bold=True)

    field_notes = [
        ("RuleName",         "Human-readable label. Appears in run summary emails."),
        ("Enabled",          "yes or no. Set to no to pause a rule without deleting it."),
        ("FromAddress",      "Exact sender email address. Leave blank to match any sender."),
        ("SubjectContains",  "Keyword to match in the subject line. Case-insensitive. Leave blank to skip."),
        ("BodyContains",     "Keyword to match in the email body. Case-insensitive. Leave blank to skip."),
        ("MinAge",           "How old an email must be before this rule applies. Use a number followed by h (hours), d (days), or m (months). Examples: 48h = 2 days, 7d = one week, 1m = one month. Cannot be less than 1h."),
        ("Action",           "'move' = move to DestinationFolder and leave it alone forever.  'pending-delete' = move to the staging folder (DestinationFolder) to await deletion. Whether that means permanent deletion or graveyard depends on NeverHardDelete in Settings."),
        ("SourceFolder",     "Folder to search for matching emails. Use 'none' to search the Inbox (default). Fill this in when chaining rules - e.g. if a previous rule moved emails to 'Promotions', set SourceFolder = Promotions on the follow-up pending-delete rule."),
        ("DestinationFolder","For 'move' rules: the folder to move emails into. For 'pending-delete' rules: should match your StagingFolderName in Settings (default: Pending Delete). Use 'none' only if Action = move and no destination is needed (unusual)."),
        ("Notes",            "Free text. Not read by the flow - just for your reference."),
    ]

    for i, (field, desc) in enumerate(field_notes, start=notes_start + 1):
        ws.cell(row=i, column=1, value=field).font = Font(bold=True, size=9)
        note_cell = ws.cell(row=i, column=2, value=desc)
        note_style(note_cell)
        note_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 28

    ws.merge_cells(f"B{notes_start}:J{notes_start}")

    ws.freeze_panes = "A2"

    return ws


# ---------------------------------------------------------------------------
# Sheet 2: SETTINGS
# ---------------------------------------------------------------------------

def build_settings_sheet(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = True

    set_col_width(ws, "A", 28)
    set_col_width(ws, "B", 20)
    set_col_width(ws, "C", 55)

    # Header row
    for col, text in enumerate(["Setting", "Value", "Description"], start=1):
        header_style(ws.cell(row=1, column=col), text)
    ws.row_dimensions[1].height = 28

    settings = [
        # ( SettingKey, DefaultValue, Description, warning_color )
        (
            "DryRunMode",
            "yes",
            "YES = log what the flow would do, but take no action. Set to 'no' only after testing.",
            True,   # warn — starts as yes intentionally
        ),
        (
            "NeverHardDelete",
            "yes",
            "YES = never permanently delete anything. Staged emails are moved to the SafeDeleteFolderName folder instead of being deleted. Recommended to keep this on. Set to 'no' only if you want actual deletion.",
            True,   # warn — safety feature, highlighted
        ),
        (
            "SafeDeleteFolderName",
            "Deleted by Script",
            "Folder where emails go when NeverHardDelete = yes. Acts as a hidden graveyard - emails pile up here instead of being permanently deleted. Create this folder in Outlook first.",
            False,
        ),
        (
            "StagingFolderName",
            "Pending Delete",
            "Name of the mailbox folder where emails wait before the hard delete flow processes them. Create this folder in Outlook first.",
            False,
        ),
        (
            "HardDeleteAfterHours",
            168,
            "How many hours a staged email sits in the Staging folder before being processed. If NeverHardDelete = yes, they are moved to SafeDeleteFolderName. If no, they are permanently deleted. 168 = 7 days.",
            False,
        ),
        (
            "GlobalMinAgeHours",
            1,
            "Safety floor. No rule can act on emails younger than this, regardless of what the Rules sheet says. Minimum: 1.",
            False,
        ),
        (
            "SendRunSummary",
            "yes",
            "Send a summary email after each flow run listing what was moved or staged.",
            False,
        ),
        (
            "SummaryEmailTo",
            "",
            "Email address to receive run summaries. Required if SendRunSummary = yes.",
            False,
        ),
        (
            "MaxEmailsPerRule",
            50,
            "Safety cap. Maximum number of emails to process per rule per run. Prevents runaway deletion if a rule is misconfigured. Increase carefully.",
            False,
        ),
        (
            "FlowVersion",
            "1.0",
            "Do not edit. Used to detect config file version mismatches.",
            False,
        ),
    ]

    for row_num, (key, value, desc, warn) in enumerate(settings, start=2):
        key_cell   = ws.cell(row=row_num, column=1, value=key)
        val_cell   = ws.cell(row=row_num, column=2, value=value)
        desc_cell  = ws.cell(row=row_num, column=3, value=desc)

        for cell in [key_cell, val_cell, desc_cell]:
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        key_cell.font = Font(bold=True, size=10)
        note_style(desc_cell)
        ws.row_dimensions[row_num].height = 32

        if warn:
            for cell in [key_cell, val_cell, desc_cell]:
                cell.fill = PatternFill("solid", fgColor=COLOR_WARNING)

    # Excel table
    last_row = 1 + len(settings)
    table = Table(displayName="SettingsTable", ref=f"A1:C{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
    )
    ws.add_table(table)

    # Warning note
    note_row = last_row + 2
    warn_cell = ws.cell(
        row=note_row,
        column=1,
        value="IMPORTANT: DryRunMode is ON by default. The flow will not move or delete anything until you change it to 'no'. Test thoroughly before disabling.",
    )
    warn_cell.font = Font(bold=True, color="C00000", size=10)
    warn_cell.fill = PatternFill("solid", fgColor=COLOR_WARNING)
    warn_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{note_row}:C{note_row}")
    ws.row_dimensions[note_row].height = 40

    ws.freeze_panes = "A2"

    return ws


# ---------------------------------------------------------------------------
# Sheet 3: RUN LOG (empty template — flow writes here)
# ---------------------------------------------------------------------------

def build_log_sheet(wb):
    ws = wb.create_sheet("Run Log")

    log_headers = [
        "Timestamp", "RuleName", "EmailSubject", "FromAddress",
        "EmailAge_Hours", "Action", "DestinationFolder", "DryRun", "Notes"
    ]
    col_widths = [22, 22, 40, 32, 16, 12, 22, 10, 40]

    for i, (h, w) in enumerate(zip(log_headers, col_widths), start=1):
        header_style(ws.cell(row=1, column=i), h)
        set_col_width(ws, get_column_letter(i), w)

    ws.row_dimensions[1].height = 28

    # Excel table — needs at least one data row or Excel strips the table on open
    table = Table(displayName="RunLogTable", ref="A1:I2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
    )
    ws.add_table(table)

    note = ws.cell(
        row=4, column=1,
        value="This sheet is written by the flow after each run. Do not edit manually. The flow appends one row per email actioned."
    )
    note_style(note)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells("A4:I4")

    ws.freeze_panes = "A2"
    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  EMAIL CLEANUP CONFIG GENERATOR")
    print("=" * 60)
    print()

    print("[1/4] Creating workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    print("      OK — blank workbook created")
    print()

    print("[2/4] Building Rules sheet...")
    build_rules_sheet(wb)
    print("      OK — RulesTable created with 4 example rules")
    print("           (3 enabled, 1 disabled)")
    print()

    print("[3/4] Building Settings sheet...")
    build_settings_sheet(wb)
    print("      OK — SettingsTable created with 10 settings")
    print("      ** DryRunMode    = yes  (flow will not take action until you change this)")
    print("      ** NeverHardDelete = yes  (nothing will ever be permanently deleted)")
    print()

    print("[4/4] Building Run Log sheet...")
    build_log_sheet(wb)
    print("      OK — RunLogTable created (empty, flow will write here)")
    print()

    print("Saving file...")
    wb.save(OUTPUT_FILE)
    print(f"DONE: {OUTPUT_FILE}")
    print()
    print("-" * 60)
    print("Next steps:")
    print("  1. Open the file and review the example rules in the Rules sheet.")
    print("  2. Replace example rules with real ones for Atlantic Logistics.")
    print("  3. Fill in SummaryEmailTo in the Settings sheet.")
    print("  4. Create a 'Pending Delete' folder in Outlook.")
    print("  5. Upload the file to OneDrive.")
    print("  6. Point both Power Automate flows at the OneDrive location.")
    print("  7. Run with DryRunMode = yes first and review the Run Log sheet.")
    print("  8. Only set DryRunMode = no after confirming the log looks correct.")
    print("-" * 60)
    print()
    try:
        input("Press Enter to close...")
    except EOFError:
        pass


if __name__ == "__main__":
    main()
